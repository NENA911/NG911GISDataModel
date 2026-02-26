# run_schema_report.py
"""
| Name:      Run Schema Report
| Purpose:   This code parses the flatfile_schema_v3.yaml file to convert
|            the YAML definition into a Microsoft Excel spreadsheet for
|            simplified review and validation.
|
| Notes:     This code is written using the default Python library for
|            ArcGIS Pro's "arcgispro-py3" conda environment. This code is not
|            dependent on ArcGIS Pro, but is part of a larger package of
|            Python scripts which are.
|
| Author:    NENA Data Structures Committee, DS-NG GIS Data Model WG
"""

import os
import yaml
from datetime import datetime
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from textwrap import dedent


CELL_FILL_PASS = PatternFill(start_color='C6EFCE', end_color='FF0000', fill_type="solid")
CELL_FILL_FAIL = PatternFill(start_color='FFC7CE', end_color='FF0000', fill_type="solid")
CELL_FILL_WARN = PatternFill(start_color='FFEB9C', end_color='FF0000', fill_type="solid")
CELL_ALIGN_CENTER = Alignment(horizontal='center')
CELL_FONT_BOLD = Font(bold=True)


def field_has_metadata_description(desc, sheet, cell):
    if len(desc) > 50:
        sheet[cell] = dedent(desc).lstrip('\n')
    else:
        sheet[cell].font = CELL_FONT_BOLD
        sheet[cell] = 'MISSING'
        sheet[cell].fill = CELL_FILL_FAIL


def is_fieldalias_in_section5(fldalias, srcalias, sheet, cell):
    sheet[cell].alignment = CELL_ALIGN_CENTER
    if fldalias == srcalias:
        sheet[cell] = 'PASS'
        sheet[cell].fill = CELL_FILL_PASS
    else:
        sheet[cell].font = CELL_FONT_BOLD
        sheet[cell] = 'FAIL'
        sheet[cell].fill = CELL_FILL_FAIL


def is_fieldname_in_section5(fldname, fldnames, sheet, cell):
    sheet[cell].alignment = CELL_ALIGN_CENTER
    if fldname in fldnames:
        sheet[cell] = 'PASS'
        sheet[cell].fill = CELL_FILL_PASS
    else:
        sheet[cell].font = CELL_FONT_BOLD
        sheet[cell] = 'FAIL'
        sheet[cell].fill = CELL_FILL_FAIL


def run_schema_report(**params):
    # ==========================================================================
    # Load YAML
    # ==========================================================================
    relative_yaml_path = os.path.join('..', '..', '..', 'schema', 'v3.0', 'flatfile_schema_v3.yaml')
    absolute_yaml_path = os.path.abspath(relative_yaml_path)
    with open(absolute_yaml_path, 'r', encoding='utf-8') as f:
        data = yaml.load(f, Loader=yaml.SafeLoader)
        DOMAINS = data['domains']
        FEATURE_CLASSES = data['feature_classes']
        FIELDS = data['fields']
        GIS_DATA_LAYERS_REGISTRY = data['gis_data_layers_registry']

    # ==========================================================================
    # Create Workbook
    # ==========================================================================
    timestamp = datetime.now().strftime('%Y%m%d%H%M')
    wb_file_path = os.path.join(params["out_path"], f'{params["report_name"]}_{timestamp}.xlsx')
    wb = Workbook()

    # ==========================================================================
    # Create Feature Class Report Worksheet
    # ==========================================================================
    ws_fcs = wb.active
    ws_fcs.title = "Feature Class Report"

    # Write the header columns
    ws_fcs.column_dimensions['A'].width = 24
    ws_fcs['A1'] = 'Feature Class'
    ws_fcs['A1'].font = CELL_FONT_BOLD
    ws_fcs.column_dimensions['B'].width = 32
    ws_fcs['B1'] = 'Field Alias'
    ws_fcs['B1'].font = CELL_FONT_BOLD
    ws_fcs['C1'] = 'Field Name'
    ws_fcs['C1'].font = CELL_FONT_BOLD
    ws_fcs.column_dimensions['C'].width = 14
    ws_fcs['D1'] = 'Type'
    ws_fcs['D1'].font = CELL_FONT_BOLD
    ws_fcs['D1'].alignment = CELL_ALIGN_CENTER
    ws_fcs.column_dimensions['D'].width = 10
    ws_fcs['E1'] = 'L'
    ws_fcs["E1"].comment = Comment('Length', '')
    ws_fcs['E1'].font = CELL_FONT_BOLD
    ws_fcs['E1'].alignment = CELL_ALIGN_CENTER
    ws_fcs.column_dimensions['E'].width = 4
    ws_fcs['F1'] = 'P'
    ws_fcs["F1"].comment = Comment('Precision', '')
    ws_fcs['F1'].font = CELL_FONT_BOLD
    ws_fcs['F1'].alignment = CELL_ALIGN_CENTER
    ws_fcs.column_dimensions['F'].width = 4
    ws_fcs['G1'] = 'S'
    ws_fcs["G1"].comment = Comment('Scale', '')
    ws_fcs['G1'].font = CELL_FONT_BOLD
    ws_fcs['G1'].alignment = CELL_ALIGN_CENTER
    ws_fcs.column_dimensions['G'].width = 4
    ws_fcs['H1'] = 'Null'
    ws_fcs["H1"].comment = Comment('Is Nullable?', '')
    ws_fcs['H1'].font = CELL_FONT_BOLD
    ws_fcs['H1'].alignment = CELL_ALIGN_CENTER
    ws_fcs.column_dimensions['H'].width = 6
    ws_fcs['I1'] = 'Reqd'
    ws_fcs["I1"].comment = Comment('Is Required?', '')
    ws_fcs['I1'].font = CELL_FONT_BOLD
    ws_fcs['I1'].alignment = CELL_ALIGN_CENTER
    ws_fcs.column_dimensions['I'].width = 6
    ws_fcs['J1'] = 'Domain'
    ws_fcs['J1'].font = CELL_FONT_BOLD
    ws_fcs.column_dimensions['J'].width = 30
    ws_fcs['K1'] = 'Default Value'
    ws_fcs['K1'].font = CELL_FONT_BOLD
    ws_fcs.column_dimensions['K'].width = 30
    ws_fcs['L1'] = 'Section #'
    ws_fcs['L1'].font = CELL_FONT_BOLD
    ws_fcs['L1'].alignment = CELL_ALIGN_CENTER
    ws_fcs.column_dimensions['L'].width = 12
    ws_fcs["L1"].comment = Comment('What is the Section Number of the field?', '')

    row = 2
    domains = [domain['domain_name'] for domain in DOMAINS]
    registry_layers = [layer['layer_name'] for layer in GIS_DATA_LAYERS_REGISTRY]
    for feature_class in FEATURE_CLASSES:
        print(feature_class["name"])
        ws_fcs[f'A{row}'] = feature_class["name"]
        if feature_class["name"] not in registry_layers:
            if feature_class["name"] == "ServiceBoundaryPolygon":
                ws_fcs[f'A{row}'].fill = CELL_FILL_WARN
                ws_fcs[f'A{row}'].comment = Comment('Template for optional ServiceBoundaryPolygon layers defined in GIS Layer Registry', '')
            else:
                ws_fcs[f'A{row}'].fill = CELL_FILL_FAIL
                ws_fcs[f'A{row}'].comment = Comment('Layer not found in GIS Layer Registry', '')
        for field in feature_class["fields"]:
            field_dict = {
                "field_name": field['field_name'],
                "field_type": field['field_type'],
                "field_precision": field['field_precision'],
                "field_scale": field['field_scale'],
                "field_length": field['field_length'],
                "field_alias": field['field_alias'],
                "field_is_nullable": field['field_is_nullable'],
                "field_is_required": field['field_is_required'],
                "field_domain": field['field_domain'],
                "default_value": field['field_default']
            }

            field_definition = None

            ws_fcs[f'B{row}'] = field_dict["field_alias"]
            ws_fcs[f'C{row}'] = field_dict["field_name"]
            if field_dict["field_name"] not in list(FIELDS.keys()):
                ws_fcs[f'C{row}'].fill = CELL_FILL_FAIL
                ws_fcs[f'C{row}'].comment = Comment('Field name not found in flatfile_schema_v3.yaml FIELDS definitions.', '')
            else:
                field_definition = FIELDS[field_dict["field_name"]]
                print(f'  {field_dict["field_name"]}')

            # field_alias test is out of order to wait for field_definition population, if exists
            if field_definition is None:
                ws_fcs[f'B{row}'].fill = CELL_FILL_FAIL
                ws_fcs[f'B{row}'].comment = Comment('Field not found in FIELDS parameters.', '')
            elif field_dict["field_alias"] != field_definition["title"]:
                ws_fcs[f'B{row}'].fill = CELL_FILL_FAIL
                ws_fcs[f'B{row}'].comment = Comment(f'{field_dict["field_alias"]} != {field_definition["title"]}', '')

            ws_fcs[f'D{row}'] = field_dict["field_type"]
            ws_fcs[f'D{row}'].alignment = CELL_ALIGN_CENTER
            ws_fcs[f'E{row}'] = field_dict["field_length"]
            ws_fcs[f'E{row}'].alignment = CELL_ALIGN_CENTER
            ws_fcs[f'F{row}'] = field_dict["field_precision"]
            ws_fcs[f'F{row}'].alignment = CELL_ALIGN_CENTER
            ws_fcs[f'G{row}'] = field_dict["field_scale"]
            ws_fcs[f'G{row}'].alignment = CELL_ALIGN_CENTER
            if field_definition is None:
                ws_fcs[f'D{row}'].fill = CELL_FILL_FAIL
                ws_fcs[f'D{row}'].comment = Comment(
                    'Field not found in FIELDS parameters.', '')
            else:
                if field_dict["field_type"] == 'DATETIME':
                    if field_definition["definition"]["type"] != 'DATETIME':
                        ws_fcs[f'D{row}'].fill = CELL_FILL_FAIL
                        ws_fcs[f'D{row}'].comment = Comment(
                            f'{field_dict["field_type"]} != {field_definition["definition"]["type"]}', '')
                elif field_dict["field_type"] == 'REAL':
                    if field_definition["definition"]["type"] != 'REAL':
                        ws_fcs[f'D{row}'].fill = CELL_FILL_FAIL
                        ws_fcs[f'D{row}'].comment = Comment(
                            f'{field_dict["field_type"]} != {field_definition["definition"]["type"]}', '')
                elif field_dict["field_type"] == 'INTEGER':
                    if field_definition["definition"]["type"] != 'INTEGER':
                        ws_fcs[f'D{row}'].fill = CELL_FILL_FAIL
                        ws_fcs[f'D{row}'].comment = Comment(
                            f'{field_dict["field_type"]} != {field_definition["definition"]["type"]}', '')
                elif field_dict["field_type"] == 'TEXT':
                    if field_definition["definition"]["type"] != 'TEXT':
                        ws_fcs[f'D{row}'].fill = CELL_FILL_FAIL
                        ws_fcs[f'D{row}'].comment = Comment(
                            f'{field_dict["field_type"]} != {field_definition["definition"]["type"]}', '')
                    if str(field_dict["field_length"]) != str(field_definition["definition"]["width"]):
                        ws_fcs[f'E{row}'].fill = CELL_FILL_FAIL
                        ws_fcs[f'E{row}'].comment = Comment(
                            f'{field_dict["field_length"]} != {field_definition["definition"]["width"]}', '')
                else:
                    ws_fcs[f'D{row}'].fill = CELL_FILL_FAIL
                    ws_fcs[f'D{row}'].comment = Comment(
                        f'Field Type: "{field_dict["field_type"]}" not recognized', '')


            ws_fcs[f'H{row}'] = field_dict["field_is_nullable"]
            ws_fcs[f'H{row}'].alignment = CELL_ALIGN_CENTER
            if str(field_dict["field_is_nullable"]).lower() == str(field_definition["definition"]["required"]).lower():
                if field_dict["field_name"] == feature_class["name"][:2]:
                    ws_fcs[f'H{row}'].fill = CELL_FILL_WARN
                    ws_fcs[f'H{row}'].comment = Comment("Special exception", '')
                else:
                    ws_fcs[f'H{row}'].fill = CELL_FILL_FAIL
                    ws_fcs[f'H{row}'].comment = Comment(
                        f'{field_dict["field_is_nullable"]} == {field_definition["definition"]["required"]}', '')


            ws_fcs[f'I{row}'] = field_dict["field_is_required"]
            ws_fcs[f'I{row}'].alignment = CELL_ALIGN_CENTER
            field_required = field_definition["definition"]["required"]
            if str(field_dict["field_is_required"]).lower() != str(field_required).lower():
                if field_dict["field_name"] == feature_class["name"][:2]:
                    ws_fcs[f'I{row}'].fill = CELL_FILL_WARN
                    ws_fcs[f'I{row}'].comment = Comment("Special exception", '')
                else:
                    ws_fcs[f'I{row}'].fill = CELL_FILL_FAIL
                    ws_fcs[f'I{row}'].comment = Comment(
                       f'{field_dict["field_is_required"]} != {field_required}', '')

            ws_fcs[f'J{row}'] = field_dict["field_domain"]
            if len(field_dict["field_domain"]) > 0 and field_dict["field_domain"] not in domains:
                ws_fcs[f'J{row}'].fill = CELL_FILL_FAIL
                ws_fcs[f'J{row}'].comment = Comment('Domain not found in flatfile_schema_v3.yaml', '')

            ws_fcs[f'K{row}'] = field_dict["default_value"]
            # Lookup field_name in nena_sta_006_v3.py FIELDS dictionary
            source_field = FIELDS.get(field_dict["field_name"])
            if not source_field:
                # If field_name is not found, mark as missing
                ws_fcs[f'L{row}'].font = CELL_FONT_BOLD
                ws_fcs[f'L{row}'].alignment = CELL_ALIGN_CENTER
                ws_fcs[f'L{row}'] = 'MISSING'
                ws_fcs[f'L{row}'].fill = CELL_FILL_FAIL
            else:
                # Add Section 5 number
                ws_fcs[f'L{row}'] = float(source_field.get("section"))
                ws_fcs[f'L{row}'].alignment = CELL_ALIGN_CENTER
            row += 1
        row += 1

    # ==========================================================================
    # Create Domains Overview
    # ==========================================================================

    ws_domains = wb.create_sheet("Domains")

    ws_domains['A1'] = 'Domain Name'
    ws_domains.column_dimensions['A'].width = 30
    ws_domains['A1'].font = CELL_FONT_BOLD

    ws_domains['B1'] = 'Domain Type'
    ws_domains.column_dimensions['B'].width = 50
    ws_domains['B1'].font = CELL_FONT_BOLD

    ws_domains['C1'] = 'Field Type'
    ws_domains.column_dimensions['C'].width = 40
    ws_domains['C1'].font = CELL_FONT_BOLD

    ws_domains['D1'] = 'Description'
    ws_domains.column_dimensions['D'].width = 50
    ws_domains['D1'].font = CELL_FONT_BOLD

    row = 2
    for domain in DOMAINS:
        ws_domains[f'A{row}'] = domain["domain_name"]
        ws_domains[f'A{row}'].alignment = Alignment(vertical='top')
        ws_domains[f'B{row}'] = domain["domain_type"]
        ws_domains[f'B{row}'].alignment = Alignment(vertical='top')
        ws_domains[f'C{row}'] = domain["field_type"]
        ws_domains[f'C{row}'].alignment = Alignment(vertical='top')
        ws_domains[f'D{row}'] = domain["domain_description"]
        ws_domains[f'D{row}'].alignment = Alignment(vertical='top', wrap_text=True)

        if domain["domain_type"] == 'CODED':
            row += 1
            ws_domains[f'B{row}'] = 'Code'
            ws_domains[f'B{row}'].font = CELL_FONT_BOLD
            ws_domains[f'C{row}'] = 'Value'
            ws_domains[f'C{row}'].font = CELL_FONT_BOLD

            if domain['values'] != None:
                for k, v in domain["values"].items():
                    row += 1
                    ws_domains[f'B{row}'] = k
                    ws_domains[f'C{row}'] = v
            else:
                row += 1
                ws_domains[f'B{row}'] = "User Defined"
                ws_domains[f'C{row}'] = "User Defined"
        elif domain["domain_type"] == 'RANGE':
            row += 1
            ws_domains[f'B{row}'] = 'Min'
            ws_domains[f'B{row}'].font = CELL_FONT_BOLD
            ws_domains[f'C{row}'] = 'Max'
            ws_domains[f'C{row}'].font = CELL_FONT_BOLD
            row += 1
            range = domain['values']
            ws_domains[f'B{row}'] = str(range['min'])
            ws_domains[f'C{row}'] = str(range['max'])

        row += 2

    # ==========================================================================
    # Create Section 4 Layer Overview
    # ==========================================================================

    ws_sec4 = wb.create_sheet("Section 4 Layer Overview")

    ws_sec4['A1'] = 'Layer Alias'
    ws_sec4.column_dimensions['A'].width = 30
    ws_sec4['A1'].font = CELL_FONT_BOLD

    ws_sec4['B1'] = 'Layer Name'
    ws_sec4.column_dimensions['B'].width = 30
    ws_sec4['B1'].font = CELL_FONT_BOLD

    ws_sec4['C1'] = 'Section #'
    ws_sec4.column_dimensions['C'].width = 10
    ws_sec4['C1'].alignment = CELL_ALIGN_CENTER
    ws_sec4['C1'].font = CELL_FONT_BOLD

    ws_sec4['D1'] = 'Geometry'
    ws_sec4.column_dimensions['D'].width = 10
    ws_sec4['D1'].alignment = CELL_ALIGN_CENTER
    ws_sec4['D1'].font = CELL_FONT_BOLD

    ws_sec4['E1'] = 'Has Z?'
    ws_sec4.column_dimensions['E'].width = 10
    ws_sec4['E1'].alignment = CELL_ALIGN_CENTER
    ws_sec4['E1'].font = CELL_FONT_BOLD

    ws_sec4['F1'] = 'Fields'
    ws_sec4.column_dimensions['F'].width = 10
    ws_sec4['F1'].alignment = CELL_ALIGN_CENTER
    ws_sec4['F1'].font = CELL_FONT_BOLD

    ws_sec4['G1'] = 'Description'
    ws_sec4.column_dimensions['G'].width = 80
    ws_sec4['G1'].font = CELL_FONT_BOLD

    ws_sec4['H1'] = 'Keywords'
    ws_sec4.column_dimensions['H'].width = 30
    ws_sec4['H1'].font = CELL_FONT_BOLD

    row = 2
    for lyr in FEATURE_CLASSES:
        ws_sec4[f'A{row}'] = lyr["alias"]
        ws_sec4[f'A{row}'].alignment = Alignment(vertical='top')
        ws_sec4[f'B{row}'] = lyr["name"]
        ws_sec4[f'B{row}'].alignment = Alignment(vertical='top')
        ws_sec4[f'C{row}'] = lyr["section"]
        ws_sec4[f'C{row}'].alignment = Alignment(vertical='top', horizontal='center')
        ws_sec4[f'D{row}'] = lyr["geometry_type"]
        ws_sec4[f'D{row}'].alignment = Alignment(vertical='top', horizontal='center')
        ws_sec4[f'E{row}'] = 'No' if lyr["has_z"] == False else 'Yes'
        ws_sec4[f'E{row}'].alignment = Alignment(vertical='top', horizontal='center')
        ws_sec4[f'F{row}'] = len(lyr["fields"])
        ws_sec4[f'F{row}'].alignment = Alignment(vertical='top', horizontal='center')
        field_has_metadata_description(
            desc=lyr["metadata"]["description"],
            sheet=ws_sec4,
            cell=f'G{row}'
        )
        ws_sec4[f'G{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws_sec4[f'H{row}'] = ", ".join(lyr["metadata"]["keywords"])
        ws_sec4[f'H{row}'].alignment = Alignment(vertical='top')
        row += 1

    # ==========================================================================
    # Create Section 5 Field Overview
    # ==========================================================================

    ws_sec5 = wb.create_sheet("Section 5 Fields Overview")

    # Write the header columns
    ws_sec5['A1'] = 'Attribute Title'
    ws_sec5.column_dimensions['A'].width = 32
    ws_sec5['A1'].font = CELL_FONT_BOLD

    ws_sec5['B1'] = 'Field Name'
    ws_sec5.column_dimensions['B'].width = 14
    ws_sec5['B1'].font = CELL_FONT_BOLD

    ws_sec5['C1'] = 'Section #'
    ws_sec5.column_dimensions['C'].width = 10
    ws_sec5['C1'].font = CELL_FONT_BOLD

    ws_sec5['D1'] = 'Description'
    ws_sec5.column_dimensions['D'].width = 80
    ws_sec5['D1'].font = CELL_FONT_BOLD

    ws_sec5['E1'] = 'Data Type'
    ws_sec5.column_dimensions['E'].width = 10
    ws_sec5['E1'].font = CELL_FONT_BOLD
    ws_sec5['E1'].alignment = CELL_ALIGN_CENTER

    ws_sec5['F1'] = 'Length'
    ws_sec5.column_dimensions['F'].width = 10
    ws_sec5['F1'].font = CELL_FONT_BOLD
    ws_sec5['F1'].alignment = CELL_ALIGN_CENTER

    ws_sec5['G1'] = 'Precision'
    ws_sec5.column_dimensions['G'].width = 10
    ws_sec5['G1'].font = CELL_FONT_BOLD
    ws_sec5['G1'].alignment = CELL_ALIGN_CENTER

    ws_sec5['H1'] = 'Scale'
    ws_sec5.column_dimensions['H'].width = 10
    ws_sec5['H1'].font = CELL_FONT_BOLD
    ws_sec5['H1'].alignment = CELL_ALIGN_CENTER

    ws_sec5['I1'] = 'Required?'
    ws_sec5.column_dimensions['I'].width = 10
    ws_sec5['I1'].font = CELL_FONT_BOLD
    ws_sec5['I1'].alignment = CELL_ALIGN_CENTER

    ws_sec5['J1'] = 'Domain?'
    ws_sec5.column_dimensions['J'].width = 10
    ws_sec5['J1'].font = CELL_FONT_BOLD
    ws_sec5['J1'].alignment = CELL_ALIGN_CENTER

    row = 2
    print("Processing Fields...")
    fgdb_fields = []
    for fc in FEATURE_CLASSES:
        for fld in fc['fields']:
            if fld['field_name'] not in fgdb_fields:
                fgdb_fields.append(fld['field_name'])
    fgdb_fields.sort()

    for fld in FIELDS.keys():
        field = FIELDS[fld]
        print(f'  {field["title"]}')
        ws_sec5[f'A{row}'] = field.get("title")
        ws_sec5[f'A{row}'].alignment = Alignment(vertical='top')
        ws_sec5[f'B{row}'] = fld
        ws_sec5[f'B{row}'].alignment = Alignment(vertical='top')
        if fld not in fgdb_fields:
            ws_sec5[f'B{row}'].fill = CELL_FILL_FAIL
            ws_sec5[f'B{row}'].comment = Comment(f'Field is not being used in flatfile_schema_v3.yaml', '')
        ws_sec5[f'C{row}'] = str(field.get("section"))
        ws_sec5[f'C{row}'].alignment = Alignment(vertical='top')
        field_has_metadata_description(
            desc=field["description"],
            sheet=ws_sec5,
            cell=f'D{row}'
        )
        ws_sec5[f'D{row}'].alignment = Alignment(vertical='top', wrap_text=True)

        field_definition = field.get("definition")
        field_type = field_definition.get("type")
        if field_type == "TEXT":
            ws_sec5[f'E{row}'] = field_type
            ws_sec5[f'E{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'F{row}'] = field_definition.get("width")
            ws_sec5[f'F{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'G{row}'] = '--'
            ws_sec5[f'G{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'H{row}'] = '--'
            ws_sec5[f'H{row}'].alignment = Alignment(vertical='top', horizontal='center')
        elif field_type == "INTEGER":
            ws_sec5[f'E{row}'] = field_type
            ws_sec5[f'E{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'F{row}'] = '--'
            ws_sec5[f'F{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'G{row}'] = '--'
            ws_sec5[f'G{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'H{row}'] = '--'
            ws_sec5[f'H{row}'].alignment = Alignment(vertical='top', horizontal='center')
        elif field_type == "REAL":
            ws_sec5[f'E{row}'] = field_type
            ws_sec5[f'E{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'F{row}'] = '--'
            ws_sec5[f'F{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'G{row}'] = field_definition.get("precision")
            ws_sec5[f'G{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'H{row}'] = field_definition.get("scale")
            ws_sec5[f'H{row}'].alignment = Alignment(vertical='top', horizontal='center')
        elif field_type == "DATETIME":
            ws_sec5[f'E{row}'] = field_type
            ws_sec5[f'E{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'F{row}'] = '--'
            ws_sec5[f'F{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'G{row}'] = '--'
            ws_sec5[f'G{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'H{row}'] = '--'
            ws_sec5[f'H{row}'].alignment = Alignment(vertical='top', horizontal='center')
        else:
            ws_sec5[f'E{row}'] = "ERROR"
            ws_sec5[f'E{row}'].alignment = Alignment(vertical='top', horizontal='center')
            ws_sec5[f'E{row}'].font = CELL_FONT_BOLD
            ws_sec5[f'E{row}'].fill = CELL_FILL_FAIL

        ws_sec5[f'I{row}'] = field_definition.get("required")
        ws_sec5[f'I{row}'].alignment = Alignment(vertical='top', horizontal='center')
        ws_sec5[f'J{row}'] = field.get("domain")
        ws_sec5[f'J{row}'].alignment = Alignment(vertical='top', horizontal='center')
        row += 1

    # ==========================================================================
    # Create GIS Layer Registry
    # ==========================================================================

    ws_reg = wb.create_sheet("GIS Layer Registry")

    ws_reg['A1'] = 'Layer Name'
    ws_reg.column_dimensions['A'].width = 40
    ws_reg['A1'].font = CELL_FONT_BOLD
    ws_reg['B1'] = 'Layer Indicator'
    ws_reg.column_dimensions['B'].width = 20
    ws_reg['B1'].font = CELL_FONT_BOLD

    fgbd_fcs = [item['name'] for item in FEATURE_CLASSES]
    optional_fcs = ['CoastGuardPolygon', 'MountainRescuePolygon', 'PoisonControlPolygon']
    row = 2
    for reg in GIS_DATA_LAYERS_REGISTRY:
        lyr_name = reg['layer_name']
        ws_reg[f'A{row}'] = lyr_name
        if lyr_name not in fgbd_fcs:
            if (lyr_name in optional_fcs) or (lyr_name.startswith('Police') and lyr_name != 'PolicePolygon') or (lyr_name.startswith('Fire') and lyr_name != 'FirePolygon') or (lyr_name.startswith('Ems') and lyr_name != 'EmsPolygon'):
                ws_reg[f'A{row}'].fill = CELL_FILL_WARN
                ws_reg[f'A{row}'].comment = Comment(f'{reg["layer_name"]} optional ServiceBoundaryPolygon layer', '')
            else:
                ws_reg[f'A{row}'].fill = CELL_FILL_FAIL
                ws_reg[f'A{row}'].comment = Comment(f'{reg["layer_name"]} not defined in flatfile_schema_v3.yaml', '')
        ws_reg[f'A{row}'].alignment = Alignment(vertical='top')
        ws_reg[f'B{row}'] = reg['layer_indicator']
        ws_reg[f'B{row}'].alignment = Alignment(vertical='top')
        row += 1

    # ==========================================================================
    # Save Workbook and close
    # ==========================================================================
    wb.save(wb_file_path)
    wb.close()


if __name__ == '__main__':
    #desktop_path = "~\\Desktop"
    desktop_path = "~\\OneDrive\\Desktop"

    params = {
        "report_name": 'NG911_GISDataModelSchemaReport_v3.0',
        "out_path": os.path.expanduser(desktop_path)
    }

    run_schema_report(**params)

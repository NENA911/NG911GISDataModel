import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from schema.nena_sta_006_v3 import FIELDS, GIS_DATA_LAYERS_REGISTRY
from schema.schema_fgdb_v3 import FEATURE_CLASSES, TABLES, DOMAINS

"""
Test - Fields are present in Section 5 - Field Names and Asoociated Attribute Data
Test - Fields in Section 5 are in tables
Test - Field types are correct
Test - Field length is correct
Test - Field required
"""

CELL_FILL_PASS = PatternFill(start_color='C6EFCE', end_color='FF0000', fill_type="solid")
CELL_FILL_FAIL = PatternFill(start_color='FFC7CE', end_color='FF0000', fill_type="solid")
CELL_ALIGN_CENTER = Alignment(horizontal='center')
CELL_FONT_BOLD = Font(bold=True)


def is_fieldname_in_section5(fldname, fldnames, sheet, cell):
    sheet[cell].alignment = CELL_ALIGN_CENTER
    if fldname in fldnames:
        sheet[cell] = 'PASS'
        sheet[cell].fill = CELL_FILL_PASS
    else:
        sheet[cell].font = CELL_FONT_BOLD
        sheet[cell] = 'FAIL'
        sheet[cell].fill = CELL_FILL_FAIL


def is_fieldalias_in_section5(fldalias, srcalias, sheet, cell):
    sheet[cell].alignment = CELL_ALIGN_CENTER
    if fldalias == srcalias:
        sheet[cell] = 'PASS'
        sheet[cell].fill = CELL_FILL_PASS
    else:
        sheet[cell].font = CELL_FONT_BOLD
        sheet[cell] = 'FAIL'
        sheet[cell].fill = CELL_FILL_FAIL


def run_schema_report(**params):

    timestamp = datetime.now().strftime('%Y%m%d%H%M')
    wb_file_path = os.path.join(params["out_path"], f'{params["report_name"]}_{timestamp}.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "Feature Class Report"

    # Write the header columns
    ws.column_dimensions['A'].width = 45
    ws['A1'] = 'Feature Class'
    ws['A1'].font = CELL_FONT_BOLD
    ws.column_dimensions['B'].width = 35
    ws['B1'] = 'Field Alias'
    ws['B1'].font = CELL_FONT_BOLD
    ws['C1'] = 'Section #'
    ws['C1'].font = CELL_FONT_BOLD
    ws['D1'] = 'Field Name?'
    ws['D1'].font = CELL_FONT_BOLD
    ws['D1'].alignment = CELL_ALIGN_CENTER
    ws['E1'] = 'Field Alias?'
    ws['E1'].font = CELL_FONT_BOLD
    ws['E1'].alignment = CELL_ALIGN_CENTER
    ws['E1'] = 'Type?'
    ws['E1'].font = CELL_FONT_BOLD
    ws['E1'].alignment = CELL_ALIGN_CENTER

    row = 2
    for feature_class in FEATURE_CLASSES:
        print(feature_class["out_name"])
        ws[f'A{row}'] = feature_class["out_name"]
        for field in feature_class["fields"]:
            field_dict = {
                "field_name": field[0],
                "field_type": field[1],
                "field_precision": field[2],
                "field_scale": field[3],
                "field_length": field[4],
                "field_alias": field[5],
                "field_is_nullable": field[6],
                "field_is_required": field[7],
                "field_domain": field[8],
                "default_value": field[9]
            }

            ws[f'B{row}'] = field_dict["field_alias"]

            source_field = FIELDS.get(field_dict["field_name"])

            if not source_field:
                ws[f'C{row}'].font = CELL_FONT_BOLD
                ws[f'C{row}'].alignment = CELL_ALIGN_CENTER
                ws[f'C{row}'] = 'MISSING'
                ws[f'C{row}'].fill = CELL_FILL_FAIL
            else:
                ws[f'C{row}'] = float(source_field.get("section"))

                ws[f'D{row}'].alignment = CELL_ALIGN_CENTER
                if field_dict["field_name"] in list(FIELDS.keys()):
                    ws[f'D{row}'] = 'PASS'
                    ws[f'D{row}'].fill = CELL_FILL_PASS
                else:
                    ws[f'D{row}'].font = CELL_FONT_BOLD
                    ws[f'D{row}'] = 'FAIL'
                    ws[f'D{row}'].fill = CELL_FILL_FAIL

                ws[f'E{row}'].alignment = CELL_ALIGN_CENTER
                if field_dict["field_alias"] == source_field['title']:
                    ws[f'E{row}'] = 'PASS'
                    ws[f'E{row}'].fill = CELL_FILL_PASS
                else:
                    ws[f'E{row}'].font = CELL_FONT_BOLD
                    ws[f'E{row}'] = 'FAIL'
                    ws[f'E{row}'].fill = CELL_FILL_FAIL

            row += 1
        row += 1

    wb.save(wb_file_path)
    wb.close()


if __name__ == '__main__':
    # desktop_path = "~\\Desktop"
    desktop_path = "~\\OneDrive\\Desktop"

    params = {
        "report_name": 'NG911_SchemaReport',
        "out_path": os.path.expanduser(desktop_path)
    }

    run_schema_report(**params)